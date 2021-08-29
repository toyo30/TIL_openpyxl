import openpyxl

target_filename = './(Yoon) 완료 68회 PreStarter 속성분석의 사본.xlsx'
wb = openpyxl.load_workbook(target_filename)
snames = wb.sheetnames

ws = wb[snames[0]]

ent_dic ={}


col_dic1 = {}
col_list = []

for sheet in snames:
    column = ws['F']
    for i in column:
        if i.value != None:
            col_list.append(i.value)

ent_dic[col_list[0]] = col_list[1:]


col_dic2 = {}
col_list2 = []
for sheet in snames:
    column = ws['P']
    for i in column:
        if i.value != None:
            col_list2.append(i.value)

ent_dic[col_list2[0]] = col_list2[1:]

print(ent_dic)


f = open('dic.txt', 'w')
f.write("{}".format(ent_dic))
f.close()



