import openpyxl

wb = openpyxl.load_workbook("./(Yoon)완료 68회 Basic - 문항속성분석.xlsx")

sheet_names = wb.sheetnames

#첫 번째 시트부터 
sheet_num = "Basic LC PartA"

ws = wb[sheet_num]

#함수 만들기 value를 리스트에 저장하는 함수

key_list = ["DistratorsVector[]", "Keywords", "Themes", "Difficulty", "PromptTense", "Word Class", "지능분류", "QuestionType", "Communication"]


dic_no = {}



col_o = []
for col in ws.iter_cols(min_row=2, min_col=15, max_row=21, max_col=19):
    for cell in col:
        if cell.value != None:
            col_o.append(cell.value)
dic_no[key_list[0]] = col_o


def find_col(row_num):
    col_num = 20
    while col_num < 30:
        key_no = 1
        value_list = []

        for col in ws.iter_cols(min_row=row_num, min_col=col_num, max_row=row_num, max_col=col_num):
            for cell in col:
                if cell.value != None:
                    value_list.append(cell.value)
                    if len(value_list) > 2:
                        dic_no[key_list[key_no]] = value_list
                    else:
                        dic_no[key_list[key_no]] = value_list[-1]
        
            col_num += 1
            key_no += 1
        
    return print("{}번 문항".format, dic_no)

row_num = 2 #2부터 시작

while row_num < 104:
    find_col(row_num)
    row_num +=20

