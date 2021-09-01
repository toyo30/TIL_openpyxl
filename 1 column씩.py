# 하나의 문제별로 딕셔너리를 만들어보기 그 다음엔 반복하기 
import openpyxl 

wb = openpyxl.load_workbook('./(Yoon)완료 68회 Basic - 문항속성분석.xlsx')
sheet_names = wb.sheetnames

#첫 번째 시트부터 
ws = wb[sheet_names[0]]


#문항번호마다 열기 
exam_no_dic = {}

# 문항번호 열기 

# for i in ws.rows

# for cell in column:
#     if i.value != None:
#         col_list2.append(i.value)


for col in ws.iter_cols(min_row=1, min_col=15, max_row=2, max_col=27):
    for cell in col:
        print(cell)



dic_no_1 = {}
col_o = [] # 각 문항별 value 값을 저장


for col in ws.iter_cols(min_row=2, min_col=15, max_row=21, max_col=19):
    for cell in col:
        if cell.value != None:
            col_o.append(cell.value)

dic_no_1["DistractorsVector[]"] = col_o

col_t = []

for col in ws.iter_cols(min_row=1, min_col=20, max_row=21, max_col=20):
    for cell in col:
        if cell.value != None:
            if len(col_t) > 1:
                col_t.append(cell.value)
                dic_no_1["DistractorsVector[]"] = col_o
            else:
                col_t.append(cell.value)
                dic_no_1[col_t[0]] = col_t[-1]

col_u = []
for col in ws.iter_cols(min_row=1, min_col=21, max_row=21, max_col=21):
    for cell in col:
        if cell.value != None:
            if len(col_u) > 2:
                col_u.append(cell.value)
                dic_no_1["DistractorsVector[]"] = col_u
            else:
                col_u.append(cell.value)
                dic_no_1[col_u[0]] = col_u[-1]


col_v = []
for col in ws.iter_cols(min_row=1, min_col=22, max_row=21, max_col=22):
    for cell in col:
        if cell.value != None:
            if len(col_v) > 2:
                col_v.append(cell.value)
                dic_no_1["DistractorsVector[]"] = col_v
            else:
                col_v.append(cell.value)
                dic_no_1[col_v[0]] = None
            
col_w = []
for col in ws.iter_cols(min_row=1, min_col=23, max_row=21, max_col=23):
    for cell in col:
        if cell.value != None:
            if len(col_w) > 2:
                col_w.append(cell.value)
                dic_no_1["DistractorsVector[]"] = col_w
            else:
                col_w.append(cell.value)
                dic_no_1[col_w[0]] = col_w[-1]

col_x = []
for col in ws.iter_cols(min_row=1, min_col=24, max_row=21, max_col=24):
    for cell in col:
        if cell.value != None:
            if len(col_x) > 2:
                col_x.append(cell.value)
                dic_no_1["DistractorsVector[]"] = col_x
            else:
                col_x.append(cell.value)
                dic_no_1[col_x[0]] = col_x[-1]

col_y = []
for col in ws.iter_cols(min_row=1, min_col=25, max_row=21, max_col=25):
    for cell in col:
        if cell.value != None:
            if len(col_y) > 2:
                col_y.append(cell.value)
                dic_no_1["DistractorsVector[]"] = col_y
            else:
                col_y.append(cell.value)
                dic_no_1[col_y[0]] = col_y[-1]

col_z = []
for col in ws.iter_cols(min_row=1, min_col=26, max_row=21, max_col=26):
    for cell in col:
        if cell.value != None:
            if len(col_z) > 2:
                col_z.append(cell.value)
                dic_no_1["DistractorsVector[]"] = col_z
            else:
                col_z.append(cell.value)
                dic_no_1[col_z[0]] = col_z[-1]

col_AA = []
for col in ws.iter_cols(min_row=1, min_col=27, max_row=21, max_col=27):
    for cell in col:
        if cell.value != None:
            if len(col_AA) > 2:
                col_AA.append(cell.value)
                dic_no_1["DistractorsVector[]"] = col_AA
            else:
                col_AA.append(cell.value)
                dic_no_1[col_AA[0]] = col_AA[-1]


col_AB = []
for col in ws.iter_cols(min_row=1, min_col=28, max_row=21, max_col=28):
    for cell in col:
        if cell.value != None:
            if len(col_AB) > 2:
                col_AB.append(cell.value)
                dic_no_1["DistractorsVector[]"] = col_AB
            else:
                col_AB.append(cell.value)
                dic_no_1[col_AB[0]] = col_AB[-1]

dic_no_1["QuestionType"] = dic_no_1.pop("유형")

print("{} 1번".format(sheet_names[0]))
print(dic_no_1)



f = open("./roti.txt", 'w')
f.write("{}".format(dic_no_1))
f.close()
