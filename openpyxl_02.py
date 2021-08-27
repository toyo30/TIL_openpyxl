import openpyxl
import os

print(os.getcwd())
os.chdir('/Users/baegmingi/pythonData/openpyxl연습/MyExcelFiles')
print(os.getcwd())
target_filename ="./myworkbook.xlsx"
dest_filename = './myworkbook_after.xlsx'

wb = openpyxl.load_workbook(target_filename) #기존 workbook을 불러온다.
print(wb.sheetnames)
wb.create_sheet(index=0, title='mysheet0') #맨 앞에 시트에 title이 mysheet 0인 것을 삽입
wb.create_sheet(title='mysheet2') #생략해서 맨 뒤에 mysheet2인 것을 삽입함

print(wb.sheetnames) #리스트 형태로 전체 시트를 불러옴

snames = wb.sheetnames #시트 네임은 wb.sheeetnames
sheet = wb['mysheet1'] #wb[snaems[1]] 1번 인덱스의 것. 모두다 mysheet에 대한 것
print(sheet['A1'].value)
print(sheet['A2'].value)
wb.save(dest_filename)



#range 사용하기 
#여러 개의 셀이 포함되어 있는 range를 사용하기